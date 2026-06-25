import streamlit as st
import pdfplumber
import spacy
import pandas as pd
import re
import io
import os
import base64
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, timezone

# ============================================================================
# 1. PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="AMR National Surveillance | USYD",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================================
# 2. MODERN BLUE THEME
# ============================================================================
BLUE_900 = "#0a2540"
BLUE_800 = "#103a6b"
BLUE_700 = "#14539a"
BLUE_600 = "#1d6fd6"
BLUE_500 = "#2f88e6"
BLUE_400 = "#5aa2ee"
BLUE_300 = "#93c2f4"
BLUE_50  = "#eef5fd"
INK      = "#0e1c2b"
MUTED    = "#5d6e80"
LINE     = "#e2eaf3"
CANVAS   = "#eaf0f7"
WHITE    = "#ffffff"

BLUE_SEQ = ["#1d6fd6", "#103a6b", "#2f88e6", "#5aa2ee", "#0a2540", "#14539a", "#93c2f4", "#3f97ea"]
SIR_CMAP = {"Susceptible": "#1aa260", "Intermediate": "#e0a400", "Resistant": "#d63a4a"}

st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

    html, body, [class*="css"], .stApp {{
        font-family: 'Inter', -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        color: {INK};
        font-size: 16px;
    }}
    .stApp {{ background-color: {CANVAS}; }}
    .block-container {{ padding-top: 1.4rem; max-width: 1340px; }}

    /* ---------- Hero ---------- */
    .amr-hero {{
        position: relative; overflow: hidden;
        background: linear-gradient(125deg, {BLUE_900} 0%, {BLUE_700} 55%, {BLUE_600} 100%);
        border-radius: 18px; padding: 2rem 2.2rem; margin-bottom: 1.6rem;
        box-shadow: 0 14px 34px rgba(10,37,64,0.28);
    }}
    .amr-hero::after {{
        content: ""; position: absolute; top: -40%; right: -8%;
        width: 360px; height: 360px; border-radius: 50%;
        background: radial-gradient(circle, rgba(255,255,255,0.12) 0%, rgba(255,255,255,0) 70%);
    }}
    .amr-hero h1 {{
        color: {WHITE}; font-size: 2.25rem; font-weight: 800; margin: 0;
        letter-spacing: -0.025em; line-height: 1.12;
    }}
    .amr-hero p {{
        color: #d3e4fb; font-size: 1.05rem; margin: 0.5rem 0 0 0; font-weight: 500;
    }}
    .amr-eyebrow {{
        display:inline-block; color:#bcd6f7; font-size:0.78rem; font-weight:700;
        letter-spacing:0.14em; text-transform:uppercase; margin-bottom:0.5rem;
    }}

    /* ---------- Section headers ---------- */
    .sec {{
        display:flex; align-items:center; gap:0.7rem;
        font-size:1.6rem; font-weight:800; color:{BLUE_800};
        letter-spacing:-0.015em; margin:0.2rem 0 1rem 0;
    }}
    .sec-bar {{ width:6px; height:28px; border-radius:6px;
        background:linear-gradient(180deg,{BLUE_500},{BLUE_700}); display:inline-block; }}
    .sec-note {{ color:{MUTED}; font-size:1rem; margin:-0.6rem 0 1.1rem 1.1rem; }}

    /* ---------- Card panels (bordered containers) ---------- */
    [data-testid="stVerticalBlockBorderWrapper"] {{
        background:{WHITE}; border:1px solid {LINE} !important; border-radius:16px;
        box-shadow:0 4px 18px rgba(14,28,43,0.06); padding:1.2rem 1.4rem;
    }}
    /* No nested block inside a card may draw its own border/background/shadow,
       otherwise Streamlit's own container border shows as a second box. */
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stVerticalBlockBorderWrapper"],
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stVerticalBlock"] {{
        border:none !important; background:transparent !important;
        box-shadow:none !important; padding:0 !important;
    }}

    /* ---------- Buttons ---------- */
    .stButton>button {{
        width:100%; border-radius:11px; height:3.2em; border:none;
        background:linear-gradient(135deg,{BLUE_600} 0%,{BLUE_500} 100%);
        color:{WHITE}; font-weight:700; font-size:1.02rem; letter-spacing:0.01em;
        box-shadow:0 6px 18px rgba(29,111,214,0.32); transition:all 0.16s ease;
    }}
    .stButton>button:hover {{
        background:linear-gradient(135deg,{BLUE_700} 0%,{BLUE_600} 100%);
        box-shadow:0 9px 24px rgba(29,111,214,0.42); transform:translateY(-1px); color:{WHITE};
    }}
    .stDownloadButton>button {{
        border-radius:11px; background:{WHITE}; color:{BLUE_700};
        border:1.6px solid {BLUE_500}; font-weight:700; height:3em;
    }}
    .stDownloadButton>button:hover {{ background:{BLUE_600}; color:{WHITE}; border-color:{BLUE_600}; }}

    /* ---------- Sidebar ---------- */
    [data-testid="stSidebar"] {{ background:{WHITE}; border-right:1px solid {LINE}; }}
    [data-testid="stSidebar"] h3 {{ color:{BLUE_800}; }}

    /* ---------- Tabs ---------- */
    .stTabs [data-baseweb="tab-list"] {{ gap:0.5rem; border-bottom:2px solid {LINE}; }}
    .stTabs [data-baseweb="tab"] {{
        font-weight:700; font-size:1.04rem; color:{MUTED}; padding:0.6rem 1.2rem;
    }}
    .stTabs [aria-selected="true"] {{ color:{BLUE_600}; }}
    .stTabs [data-baseweb="tab-highlight"] {{ background:{BLUE_600}; height:3px; border-radius:3px; }}

    /* ---------- Metric cards ---------- */
    [data-testid="stMetric"] {{
        background:linear-gradient(180deg,{WHITE} 0%,{BLUE_50} 130%);
        border:1px solid {LINE}; border-left:5px solid {BLUE_600};
        border-radius:14px; padding:1.25rem 1.45rem; box-shadow:0 3px 12px rgba(14,28,43,0.05);
    }}
    [data-testid="stMetricValue"] {{ color:{BLUE_700}; font-weight:900; font-size:2.7rem; }}
    [data-testid="stMetricLabel"] {{ color:{MUTED}; font-weight:600; font-size:1.02rem; }}

    /* ---------- Uploaders (blend into the section card) ---------- */
    [data-testid="stFileUploader"] {{ background:transparent; border:none; padding:0; }}
    [data-testid="stFileUploader"] label {{ font-weight:600; color:{BLUE_800}; font-size:0.9rem; }}
    [data-testid="stFileUploaderDropzone"] {{
        background:{BLUE_50}; border:1.4px dashed {BLUE_300}; border-radius:11px;
        padding:0.5rem 0.9rem; min-height:auto;
    }}
    [data-testid="stFileUploaderDropzone"] section {{ background:transparent; }}
    [data-testid="stFileUploaderDropzoneInstructions"] span,
    [data-testid="stFileUploaderDropzoneInstructions"] small {{ font-size:0.74rem; }}
    [data-testid="stFileUploaderDropzone"] button {{ font-size:0.78rem; padding:0.2rem 0.7rem; }}

    /* ---------- Dataframe (sits inside a card already) ---------- */
    [data-testid="stDataFrame"] {{ border-radius:10px; overflow:hidden; border:none; }}

    .legend-chip {{ display:inline-block; width:14px; height:14px; border-radius:4px;
        margin-right:8px; vertical-align:middle; }}
    .skip-box {{
        background:{BLUE_50}; border:1px solid {LINE}; border-left:4px solid {BLUE_400};
        border-radius:12px; padding:0.9rem 1.1rem;
    }}
    .skip-box b {{ color:{BLUE_800}; }}
    .skip-box ol {{ margin:0.5rem 0 0 1.1rem; color:{INK}; }}
    hr {{ border-color:{LINE}; }}
    </style>
""", unsafe_allow_html=True)


def section(title, note=None):
    st.markdown(f"<div class='sec'><span class='sec-bar'></span>{title}</div>", unsafe_allow_html=True)
    if note:
        st.markdown(f"<div class='sec-note'>{note}</div>", unsafe_allow_html=True)


def style_fig(fig, height, bottom=40):
    fig.update_layout(
        height=height, template="simple_white", plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=INK, size=15.5, family="Inter"), margin=dict(b=bottom, t=24, l=4, r=4),
        hoverlabel=dict(font=dict(size=16, family="Inter"), bgcolor="white", bordercolor=LINE),
    )
    fig.update_xaxes(showline=True, linewidth=1.4, linecolor=LINE, tickfont=dict(size=14))
    fig.update_yaxes(showline=True, linewidth=1.4, linecolor=LINE, gridcolor=LINE, tickfont=dict(size=14))
    return fig


@st.cache_data
def logo_data_uri():
    """Return a base64 data URI for the white USYD logo if it sits next to app.py."""
    here = os.path.dirname(os.path.abspath(__file__))
    for name in ("usyd_logo_white.png", "assets/usyd_logo_white.png"):
        path = os.path.join(here, name)
        if os.path.exists(path):
            with open(path, "rb") as fh:
                return "data:image/png;base64," + base64.b64encode(fh.read()).decode()
    return None


# ============================================================================
# 3. CONFIGURATION
# ============================================================================
AST_FIRST_ABX_COL = 13
AST_CP_COL = 1
AST_ISOLATE_COL = 10
AST_SKIP_SHEETS = {"MALDI ID NAME LIST"}

CANON_ABBREVS = [
    "AM10", "AMC30", "CZN 30", "CL30", "CAZ30", "CVN30", "DO30", "SXT", "ENR5",
    "MAR", "GM 10", "C30", "P10", "F/M300", "N30", "OX1", "FOX30", "CC2", "E15",
    "TIC75", "TIM85", "S5", "CN 120", "RD5", "FA10", "AN30", "IPM10", "VA30",
]

# Columns sourced from the PDF report.
PDF_COLUMNS = [
    "Arrival Date", "Report Date", "Lab Reference", "Species", "Breed", "Age",
    "Sex", "Neutered", "Sample Type", "Site", "Sample Site (Detailed)", "Purity", "Isolate",
]
# Final column order. Clinic and MALDI Score come from the AST sheet.
OUTPUT_COLUMNS = [
    "Arrival Date", "Report Date", "Lab Reference", "Clinic", "Species", "Breed", "Age",
    "Sex", "Neutered", "Sample Type", "Site", "Sample Site (Detailed)", "Purity",
    "Isolate", "MALDI Score",
] + CANON_ABBREVS

GREEN = "FFC6EFCE"; YELLOW = "FFFFEB9C"; RED = "FFFFC7CE"
ABX_HEADER_COLORS = {
    "AM10": GREEN, "CZN 30": GREEN, "CL30": GREEN, "DO30": GREEN, "SXT": GREEN,
    "C30": GREEN, "P10": GREEN, "CC2": GREEN, "E15": GREEN, "FA10": GREEN,
    "AMC30": YELLOW, "GM 10": YELLOW, "N30": YELLOW, "TIM85": YELLOW, "CN 120": YELLOW,
    "ENR5": RED, "MAR": RED, "CVN30": RED, "OX1": RED, "F/M300": RED,
    "RD5": RED, "AN30": RED, "IPM10": RED, "VA30": RED,
}

# Valid antibiotic-cell values. Anything else is treated as a likely typo and
# highlighted bright red so it can be checked.
VALID_SIR = {"S", "I", "R", "INTR"}


def cell_css(v):
    """Cell styling for the antibiotic columns only (applied via subset)."""
    if v == "S":
        return "background-color: #C6EFCE"
    if v == "I":
        return "background-color: #FFEB9C"
    if v == "R":
        return "background-color: #FFC7CE"
    if v == "INTR":
        return "background-color: #DCE1E8"
    if v is None or v == "" or v == "NA" or (isinstance(v, float) and pd.isna(v)):
        return ""
    # Unexpected value (typo / bad entry): flag loudly.
    return "background-color: #FF2D2D; color: white; font-weight: 700"

# ============================================================================
# 4. MODEL LOADING
# ============================================================================
@st.cache_resource
def load_nlp():
    return spacy.load("en_core_web_sm")
nlp = load_nlp()

# ============================================================================
# 5. HELPERS
# ============================================================================
def normalize_cp(value):
    if value is None:
        return None
    m = re.search(r'(\d{2}-\d{4,5})', str(value))
    return m.group(1) if m else None


def redact_text(text):
    if not isinstance(text, str):
        return "NA"
    doc = nlp(text)
    for ent in doc.ents:
        if ent.label_ in ["PERSON", "GPE", "LOC"]:
            text = text.replace(ent.text, "[REDACTED]")
    return text


def standardize_age(age_string):
    if not age_string:
        return "NA"
    years, months = 0, 0
    ym = re.search(r'(\d+)\s*(y|year|years)', age_string, re.IGNORECASE)
    if ym:
        years = int(ym.group(1))
    mm = re.search(r'(\d+)\s*(m|month|months)', age_string, re.IGNORECASE)
    if mm:
        months = int(mm.group(1))
    return f"{years}Y {months}M"


def standardize_date(date_string):
    if not date_string or date_string == "NA":
        return "NA"
    try:
        clean = re.sub(r'^[a-zA-Z]+,\s*', '', date_string)
        clean = re.sub(r'\s+\d{1,2}:\d{2}\s+[APMpm]{2}$', '', clean)
        return datetime.strptime(clean.strip(), "%d %B %Y").strftime("%Y%m%d")
    except Exception:
        try:
            dt = pd.to_datetime(date_string, errors='coerce')
            if pd.notna(dt):
                return dt.strftime("%Y%m%d")
        except Exception:
            pass
    return date_string


def clean_boilerplate(text):
    junk = ["SYDNEY SCHOOL", "FACULTY OF VET", "PATHOLOGY DIAGNOSTIC", "UNIVERSITY OF SYDNEY",
            "CRICOS", "ABN 15", "FINAL REPORT"]
    out = []
    for line in text.split('\n'):
        l = line.strip()
        if not l:
            continue
        if any(j.lower() in l.lower() for j in junk):
            continue
        if re.search(r'Page:\s*\d+|T[: \s]*02\s*9351|date:|Ref:', l, re.IGNORECASE):
            continue
        out.append(l)
    return "\n".join(out)


def clean_isolate_name(name):
    if pd.isna(name):
        return "NA"
    name = str(name)
    name = re.sub(r'^\d+[\.\)]\s*', '', name)
    name = re.sub(r'^(?:Heavy|Moderate|Light|Scanty|Profuse|Abundant|Mixed)\s*growth\s*(?:of\s*)?(?:[-–—]\s*)?', '', name, flags=re.IGNORECASE)
    name = re.sub(r'^\d+[\.\)]\s*', '', name)
    name = re.sub(r'^[-–—\s]+', '', name)
    name = " ".join(name.split()).capitalize()
    return name if name else "NA"


# ============================================================================
# 6. PDF METADATA EXTRACTION  (everything except antibiotic results)
# ============================================================================
def parse_pdf_metadata(file_object):
    records = []
    with pdfplumber.open(file_object) as pdf:
        raw_text = "".join((page.extract_text() or "") + "\n" for page in pdf.pages)

    cp_key = normalize_cp(re.search(r'Our Ref:\s*(?:CP\s*)?(\d{2}-\d{4,5})', raw_text, re.IGNORECASE))
    lab_ref_val = f"CP {cp_key}" if cp_key else "NA"

    rd = re.search(r'Report date:\s*(.*?)(?=\s*Page:|\n)', raw_text, re.IGNORECASE)
    report_date_val = standardize_date(rd.group(1).strip() if rd else "NA")

    ad = re.search(r'Arrival date:\s*(.*?)(?=\s*\]|\s*Page:|\n)', raw_text, re.IGNORECASE)
    arrival_date_val = standardize_date(ad.group(1).strip() if ad else "NA")

    sb = re.search(r'(Canine|Feline)[\s\-]+([a-zA-Z\s\-]+?)(?=\s*(?:\n|Male|Female|\d+\s*Years?|Our Ref|$))', raw_text, re.IGNORECASE)
    species_val = sb.group(1).strip().capitalize() if sb else "NA"
    breed_val = sb.group(2).strip(" -") if sb else "NA"

    age = re.search(r'(\d+\s*(?:Years?|Months?|Weeks?))', raw_text, re.IGNORECASE)
    age_val = standardize_age(age.group(1)) if age else "NA"

    g = re.search(r'(Male Neutered|Female Spayed|Male|Female)', raw_text, re.IGNORECASE)
    if g and "Neutered" in g.group(1):
        sex_val, neutered_val = "Male", "Yes"
    elif g and "Spayed" in g.group(1):
        sex_val, neutered_val = "Female", "Yes"
    elif g:
        sex_val, neutered_val = g.group(1).capitalize(), "No"
    else:
        sex_val, neutered_val = "NA", "NA"

    clean_text = clean_boilerplate(raw_text)
    sample_blocks = re.split(r'^SAMPLE(?:\s+\d+)?\s*$', clean_text, flags=re.IGNORECASE | re.MULTILINE)
    blocks = sample_blocks[1:] if len(sample_blocks) > 1 else [clean_text]

    for block in blocks:
        sample_line = block.strip().split('\n')[0].strip()
        sample_type_val, sample_site_val = (sample_line.split(':', 1) + ["NA"])[:2] if ':' in sample_line else (sample_line, "NA")
        if sample_site_val == "NA":
            sf = re.search(r'(Swab|Urine|Tissue|Fluid|Implant):\s*(.+)', block, re.IGNORECASE)
            if sf:
                sample_type_val, sample_site_val = sf.groups()

        sample_site_val = redact_text(sample_site_val).strip()
        sample_site_detailed_val = "NA"
        if sample_site_val and sample_site_val != "NA":
            pm = re.search(r'^(.*?)\s*\((.*?)\)', sample_site_val)
            if pm:
                sample_site_val = pm.group(1).strip()
                dp = pm.group(2).strip()
                if dp:
                    sample_site_detailed_val = dp[0].upper() + dp[1:]
            else:
                ss = re.split(r'[,/:\-;|]', sample_site_val, maxsplit=1)
                if len(ss) > 1:
                    sample_site_val = ss[0].strip()
                    dp = re.sub(r'[\)\]\}]+$', '', ss[1].strip()).strip()
                    if dp:
                        sample_site_detailed_val = dp[0].upper() + dp[1:]
            sample_site_val = (sample_site_val[0].upper() + sample_site_val[1:]) if sample_site_val else "NA"

        names = []
        for m in re.finditer(r'([A-Z][a-z]+\s+(?:sp\.|spp\.|[a-z]+))\s*(?:\n\s*)*SUSCEPTIBILITY', block):
            names.append(m.group(1))
        for m in re.finditer(r'MALDI-TOF Identification\s*\n+\s*(?:\d+\.\s*(?:(?:Heavy|Moderate|Light|Scanty|Profuse|Abundant|Mixed)\s*growth\s*(?:of\s*)?(?:[-–—]\s*)?)?)?([A-Z][a-z]+\s+(?:sp\.|spp\.|[a-z]+))', block, re.IGNORECASE):
            names.append(m.group(1))
        for m in re.finditer(r'\b[1-9]\.\s+([A-Z][a-z]+\s+(?:sp\.|spp\.|[a-z]+))', block, re.IGNORECASE):
            names.append(m.group(1))

        unique_ids = sorted(set(names), key=lambda x: block.find(x))
        purity = "Mixed" if len([u for u in unique_ids if clean_isolate_name(u) != "NA"]) > 1 else "Pure"

        for iso in unique_ids:
            iso_clean = clean_isolate_name(iso)
            if iso_clean == "NA":
                continue
            records.append({
                "Arrival Date": arrival_date_val,
                "Report Date": report_date_val,
                "Lab Reference": lab_ref_val,
                "Species": species_val,
                "Breed": breed_val,
                "Age": age_val,
                "Sex": sex_val,
                "Neutered": neutered_val,
                "Sample Type": sample_type_val.strip(),
                "Site": sample_site_val,
                "Sample Site (Detailed)": sample_site_detailed_val,
                "Purity": purity,
                "Isolate": iso_clean,
                "_cp_key": cp_key,
                "_isolate_key": iso_clean.strip().lower(),
            })

    return records


# ============================================================================
# 7. AST ANTIBIOTIC LOOKUP
# ============================================================================
AST_CLINIC_COL = 3
AST_MALDI_SCORE_COL = 11


def resolve_cell(meas, sir):
    """Resolve one antibiotic cell from its (measurement, S/I/R) pair.
    If the measurement cell reads INTR (intrinsic resistance), record INTR.
    Otherwise take the S/I/R interpretation. Unexpected text is kept verbatim
    so the styler can flag it as a likely typo."""
    if meas is not None and str(meas).strip().upper() == "INTR":
        return "INTR"
    if sir is None:
        return "NA"
    s = str(sir).strip()
    if s.upper() in ("", "N/A", "NA"):
        return "NA"
    if s.upper() in ("S", "I", "R"):
        return s.upper()
    return s  # unexpected -> kept, highlighted red downstream


def build_ast_lookup(file_object):
    wb = openpyxl.load_workbook(file_object, read_only=True, data_only=True)
    meas_cols = {abx: AST_FIRST_ABX_COL + 2 * i for i, abx in enumerate(CANON_ABBREVS)}
    sir_cols = {abx: AST_FIRST_ABX_COL + 1 + 2 * i for i, abx in enumerate(CANON_ABBREVS)}
    lookup = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in AST_SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        if not header or len(header) <= AST_CP_COL or str(header[AST_CP_COL]).strip().upper() != "CP NUMBER":
            continue
        for row in rows[1:]:
            if len(row) <= AST_CP_COL or not row[AST_CP_COL]:
                continue
            cp_key = normalize_cp(row[AST_CP_COL])
            iso = row[AST_ISOLATE_COL] if AST_ISOLATE_COL < len(row) else None
            if not cp_key or not iso:
                continue
            key = (cp_key, str(iso).strip().lower())
            maldi_raw = row[AST_MALDI_SCORE_COL] if AST_MALDI_SCORE_COL < len(row) else None
            try:
                maldi_val = f"{float(maldi_raw):.2f}" if maldi_raw not in (None, "") else "NA"
            except (ValueError, TypeError):
                maldi_val = str(maldi_raw)
            entry = {
                "Clinic": row[AST_CLINIC_COL] if AST_CLINIC_COL < len(row) and row[AST_CLINIC_COL] is not None else "NA",
                "MALDI Score": maldi_val,
            }
            for abx in CANON_ABBREVS:
                mi, si = meas_cols[abx], sir_cols[abx]
                meas = row[mi] if mi < len(row) else None
                sir = row[si] if si < len(row) else None
                entry[abx] = resolve_cell(meas, sir)
            lookup.setdefault(key, entry)
    return lookup


# ============================================================================
# 8. MERGE
# ============================================================================
def build_dataframe(pdf_records, ast_lookup):
    rows, skipped = [], []
    for rec in pdf_records:
        key = (rec["_cp_key"], rec["_isolate_key"])
        if key not in ast_lookup:
            skipped.append(f'{rec["Lab Reference"]}: {rec["Isolate"]}')
            continue
        vals = ast_lookup[key]
        merged = {col: rec[col] for col in PDF_COLUMNS}
        merged["Clinic"] = vals.get("Clinic", "NA")
        merged["MALDI Score"] = vals.get("MALDI Score", "NA")
        for abx in CANON_ABBREVS:
            merged[abx] = vals.get(abx, "NA")
        rows.append(merged)
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    # Same case + same organism + identical susceptibility profile -> report once.
    # If any antibiotic differs, both rows are kept.
    if not df.empty:
        df = df.drop_duplicates(subset=["Lab Reference", "Isolate"] + CANON_ABBREVS,
                                keep="first").reset_index(drop=True)
    return df, skipped


def build_excel(df):
    styled = df.style.map(cell_css, subset=CANON_ABBREVS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        styled.to_excel(writer, index=False, sheet_name="AMR Surveillance")
        ws = writer.sheets["AMR Surveillance"]
        for col_num, col_name in enumerate(df.columns, 1):
            if col_name in ABX_HEADER_COLORS:
                fill = ABX_HEADER_COLORS[col_name]
                ws.cell(1, col_num).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    return buf.getvalue()


# ============================================================================
# 9. SIDEBAR
# ============================================================================
with st.sidebar:
    st.markdown(f"<h3 style='margin-bottom:0;color:{BLUE_800};'>USYD · Veterinary Pathology</h3>"
                f"<p style='color:{MUTED};font-size:0.9rem;margin-top:0.2rem;'>AMR Surveillance Pipeline</p>",
                unsafe_allow_html=True)
    st.divider()
    st.markdown("**Workflow**")
    st.markdown(
        "1. Upload the AST LOGGING sheet(s)\n"
        "2. Add the matching PDF report(s)\n"
        "3. Process: match on CP number and isolate\n"
        "4. Download the master sheet")
    st.divider()
    st.markdown("**Result key**")
    st.markdown(
        f"<span class='legend-chip' style='background:#C6EFCE;'></span>Susceptible (S)<br>"
        f"<span class='legend-chip' style='background:#FFEB9C;'></span>Intermediate (I)<br>"
        f"<span class='legend-chip' style='background:#FFC7CE;'></span>Resistant (R)",
        unsafe_allow_html=True)
    st.divider()
    st.caption("Metadata is read from the PDF report. Antibiotic results are read from the AST sheet.")

# ============================================================================
# 10. MAIN
# ============================================================================
_logo = logo_data_uri()
_logo_html = (f"<img src='{_logo}' alt='The University of Sydney' "
              f"style='height:50px;margin-bottom:1rem;display:block;'/>") if _logo else \
             "<span class='amr-eyebrow'>Sydney School of Veterinary Science</span>"

st.markdown(
    "<div class='amr-hero'>"
    f"{_logo_html}"
    "<h1>AMR National Surveillance Pipeline</h1>"
    "<p>Automated antimicrobial susceptibility surveillance &middot; report &amp; AST integration</p>"
    "</div>",
    unsafe_allow_html=True)

tab1, tab2 = st.tabs(["Data Processing", "Analytics"])

with tab1:
    with st.container(border=True):
        section("Upload sources", "Drop in the AST LOGGING sheet(s) and the matching PDF report(s).")
        c1, c2 = st.columns(2)
        with c1:
            ast_file = st.file_uploader("AST LOGGING sheet(s) (Excel)", type=["xlsx"])
        with c2:
            pdf_files = st.file_uploader("PDF report(s)", type=["pdf"], accept_multiple_files=True)
        run = st.button("Process & Synchronise")

    if run:
        if not ast_file:
            st.error("Please upload the AST LOGGING sheet.")
        elif not pdf_files:
            st.error("Please upload at least one PDF report.")
        else:
            ast_lookup = build_ast_lookup(ast_file)
            pdf_records, bad = [], []
            total = len(pdf_files)
            pb = st.progress(0.0, text=f"Reading reports (0/{total})…")
            for i, f in enumerate(pdf_files):
                pb.progress(i / total, text=f"Reading ({i + 1}/{total}): {f.name}")
                try:
                    pdf_records.extend(parse_pdf_metadata(f))
                except Exception as e:
                    bad.append(f"{f.name}: {e}")
            pb.progress(1.0, text="Done")

            final_df, skipped = build_dataframe(pdf_records, ast_lookup)

            if final_df.empty:
                st.warning("No isolates matched between the PDF(s) and the AST sheet. Check the CP numbers and isolate names line up.")
            else:
                st.session_state['processed_data'] = final_df
                st.session_state['skipped'] = skipped
                st.session_state['bad'] = bad

    if 'processed_data' in st.session_state:
        final_df = st.session_state['processed_data']
        with st.container(border=True):
            section("Master dataset")
            preview = final_df.copy()
            preview.index = range(1, len(preview) + 1)
            st.dataframe(preview.style.map(cell_css, subset=CANON_ABBREVS), use_container_width=True)
            aus_time = datetime.now(timezone.utc) + timedelta(hours=10)
            fname = f"AMR_Surveillance_{aus_time.strftime('%Y%m%d')}.xlsx"
            st.download_button("Download master Excel", build_excel(final_df), fname,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        skipped = st.session_state.get('skipped')
        if skipped:
            items = "".join(f"<li>{s}</li>" for s in skipped)
            st.markdown(
                f"<div class='skip-box'><b>Isolates with no matching AST row (not included)</b>"
                f"<ol>{items}</ol></div>", unsafe_allow_html=True)
        if st.session_state.get('bad'):
            for b in st.session_state['bad']:
                st.error(b)

with tab2:
    if 'processed_data' not in st.session_state:
        st.info("Process data in the first tab to unlock analytics.")
    else:
        df = st.session_state['processed_data'].copy()
        clean = df[~df["Isolate"].isin(["nan", "NA", "Na", ""])]

        # ---- KPI strip ----
        m1, m2, m3 = st.columns(3)
        m1.metric("Total isolates", len(clean))
        m2.metric("Unique clinical cases", clean["Lab Reference"].nunique())
        m3.metric("Unique bacteria types", clean["Isolate"].nunique())
        st.write("")

        # ---- Bacterial species distribution (horizontal, sorted, labelled) ----
        with st.container(border=True):
            section("Bacterial species distribution",
                    "Number of isolates identified for each organism across all processed reports.")
            counts = clean["Isolate"].value_counts().sort_values(ascending=True)
            y_cats, x_vals = counts.index.tolist(), [int(v) for v in counts.values]
            fig = go.Figure(go.Bar(
                orientation="h", y=y_cats, x=x_vals, marker_color=BLUE_600, marker_line_width=0,
                text=x_vals, textposition="outside", textfont=dict(size=14, color=INK), cliponaxis=False,
                hovertemplate="<b>%{y}</b><br>Isolates: %{x}<extra></extra>"))
            style_fig(fig, max(340, 38 * len(y_cats)), bottom=30)
            fig.update_layout(xaxis_title="<b>Number of isolates</b>", yaxis_title="")
            fig.update_xaxes(range=[0, (max(x_vals) if x_vals else 1) * 1.18])
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("Show verification table"):
                tbl = pd.DataFrame({"Bacterial species": y_cats[::-1], "Isolates": x_vals[::-1]})
                tbl.index = range(1, len(tbl) + 1)
                st.dataframe(tbl, use_container_width=True)

        st.write("")

        # ---- Resistance profile (one sleek sorted 100% stacked bar, n shown inline) ----
        sir_cols = [c for c in CANON_ABBREVS if c in df.columns and df[c].isin(['S', 'I', 'R']).any()]
        if sir_cols:
            long = df[["Isolate"] + sir_cols].melt(id_vars="Isolate", var_name="ABx", value_name="Res")
            long = long[long["Res"].isin(["S", "I", "R"])]
            long["Res"] = long["Res"].map({'S': 'Susceptible', 'I': 'Intermediate', 'R': 'Resistant'})
            ct = long.groupby(['ABx', 'Res']).size().unstack(fill_value=0)
            for c in ['Resistant', 'Intermediate', 'Susceptible']:
                if c not in ct.columns:
                    ct[c] = 0
            ct['n'] = ct[['Resistant', 'Intermediate', 'Susceptible']].sum(axis=1)
            ct = ct[ct['n'] > 0]
            pct = ct[['Resistant', 'Intermediate', 'Susceptible']].div(ct['n'], axis=0) * 100
            order_cats = pct['Resistant'].sort_values(ascending=True).index.tolist()
            labels = [f"{a}  ·  n={int(ct.loc[a, 'n'])}" for a in order_cats]

            # Which organisms sit behind each (antibiotic, result) cell, for the hover.
            org_map = {}
            for (a, r), grp in long.groupby(['ABx', 'Res'])['Isolate']:
                vc = grp.value_counts()
                org_map[(a, r)] = "<br>".join(f"{name} ×{c}" if c > 1 else name for name, c in vc.items())

            with st.container(border=True):
                section("Resistance profile by antibiotic",
                        "Each bar is 100% of the isolates tested for that drug, split into S / I / R and sorted by resistance. "
                        "n = number of isolates tested (small n = interpret with caution). Hover a segment to see which organisms it is.")
                figR = go.Figure()
                for res in ["Resistant", "Intermediate", "Susceptible"]:
                    cdata = [[int(ct.loc[a, res]), org_map.get((a, res), "")] for a in order_cats]
                    figR.add_bar(orientation="h", y=labels,
                        x=[round(pct.loc[a, res], 1) for a in order_cats], name=res,
                        marker_color=SIR_CMAP[res], marker_line=dict(color="white", width=1),
                        customdata=cdata,
                        hovertemplate="<b>%{y}</b><br>" + res + ": %{x:.0f}% (%{customdata[0]} isolates)"
                                      "<br>%{customdata[1]}<extra></extra>")
                style_fig(figR, max(420, 27 * len(order_cats)), bottom=40)
                figR.update_layout(barmode="stack",
                    legend=dict(orientation="h", yanchor="bottom", y=1.01, x=0, title="",
                                font=dict(size=14)),
                    xaxis_title="<b>Percentage of isolates tested (%)</b>", yaxis_title="")
                figR.update_xaxes(range=[0, 100], ticksuffix="%")
                st.plotly_chart(figR, use_container_width=True)

            st.write("")

        # ---- Sample site (horizontal, sorted, labelled) ----
        with st.container(border=True):
            section("Sample site distribution")
            clean_sites = df[~df["Site"].isin(["nan", "NA", "Na", ""])]
            s_counts = clean_sites["Site"].value_counts().sort_values(ascending=True)
            ys, xs = s_counts.index.tolist(), [int(v) for v in s_counts.values]
            figs = go.Figure(go.Bar(
                orientation="h", y=ys, x=xs, marker_color=BLUE_500, marker_line_width=0,
                text=xs, textposition="outside", textfont=dict(size=14, color=INK), cliponaxis=False,
                hovertemplate="<b>%{y}</b><br>Count: %{x}<extra></extra>"))
            style_fig(figs, max(300, 42 * len(ys)), bottom=30)
            figs.update_layout(xaxis_title="<b>Number of isolates</b>", yaxis_title="")
            figs.update_xaxes(range=[0, (max(xs) if xs else 1) * 1.2])
            st.plotly_chart(figs, use_container_width=True)

        st.write("")

        # ---- Breed prevalence (donut circles) at the end ----
        with st.container(border=True):
            section("Breed prevalence by host species",
                    "One count per unique clinical case.")
            unique_demo = clean.drop_duplicates(subset=['Lab Reference'])
            host_species = [s for s in unique_demo["Species"].unique() if s not in ("NA", "nan")]
            if host_species:
                dcols = st.columns(len(host_species))
                for col, host in zip(dcols, host_species):
                    sub = unique_demo[unique_demo["Species"] == host]
                    with col:
                        figd = px.pie(sub, names="Breed", hole=0.58, color_discrete_sequence=BLUE_SEQ)
                        figd.update_traces(textposition="outside", textinfo="label+percent",
                            textfont=dict(size=14, family="Inter"),
                            marker=dict(line=dict(color="white", width=2)),
                            hovertemplate="<b>%{label}</b><br>Cases: %{value}<extra></extra>")
                        figd.update_layout(height=440, template="simple_white", paper_bgcolor="rgba(0,0,0,0)",
                            title=dict(text=f"<b>{host} breeds</b>", x=0.5, xanchor="center",
                                       font=dict(size=19, color=BLUE_800)),
                            font=dict(color=INK, size=14.5, family="Inter"),
                            hoverlabel=dict(font=dict(size=16, family="Inter"), bgcolor="white", bordercolor=LINE),
                            showlegend=False, margin=dict(t=64, b=34, l=24, r=24),
                            annotations=[dict(text=f"<b>{len(sub)}</b><br>cases", x=0.5, y=0.5,
                                              font=dict(size=17, color=BLUE_700), showarrow=False)])
                        st.plotly_chart(figd, use_container_width=True)
            else:
                st.info("No host-species demographic data available.")
